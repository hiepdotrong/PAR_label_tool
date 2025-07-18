import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from PIL import Image, ImageTk
import os
import glob
from openpyxl import Workbook, load_workbook

# Resample cho Pillow >=10
try:
    resample_method = Image.Resampling.LANCZOS
except AttributeError:
    resample_method = Image.ANTIALIAS

# Dịch sang tiếng Anh
translation_dict = {
    "nam": "male", "nữ": "female", "không rõ": "unknown", "không có": "none",
    "1-5": "1-5", "5-10": "5-10", "10s": "10s", "20s": "20s",
    "30s": "30s", "40s": "40s", "50s": "50s", "60s": "60s", "70s": "70s", "80s": "80s",

    "gầy": "thin", "bình thường": "normal", "mập": "fat",

    "đen": "black", "nâu": "brown", "vàng": "yellow", "trắng": "white", "xanh": "blue", "đỏ": "red", "xám": "gray",
    "xanh lá": "green", "hồng": "pink", "tím": "purple",
    
    "ngắn": "short", "trung bình": "medium", "dài": "long", # Độ dài tóc 

    "mũ bảo hiểm": "helmet", "mũ công nhân": "worker hat", "mũ lưỡi trai": "baseball cap", 
    "nón lá": "conical hat", "mũ tai bèo": "bucket hat", "mũ nồi": "beret", # Loại mũ

    "áo phông": "t-shirt", "áo sơ mi": "button-up", "áo khoác": "jacket",
    "áo phông nữ": "blouse", "áo vest": "suit", "váy": "dress", "áo polo": "polo",
    "áo khoác dày": "coat", "măng tô": "trench coat", "hoodie": "hoodie", # Loại áo

    "ngắn tay": "short sleeve", "dài tay": "long sleeve", # Loại tay áo

    "sọc ngang": "horizontal striped", "sọc dọc": "vertical striped", "trơn": "plain",
    "in chữ": "letter printed", "in hình": "printed", "logo": "logo", "kẻ caro": "checked",
    "đốm": "dotted", "màu sặc sỡ": "floral", "có vùng khác màu": "color block", # Họa tiết áo/quần

    "quần bò": "jeans", "quần short/quần đùi": "shorts", "quần thô túi hộp": "cargo", "quần vải bó chân": "leggings",
    "quần vải chỉ qua đầu gối": "capri", "quần vải thể thao": "track pants", "quần âu": "trousers", "chân váy": "skirt",
    "váy": "dress", "quần cạp cao có dây đeo vai": "overalls",  # Loại quần

    "giày thể thao": "sneakers", "dép": "sandals", "tông": "flip-flops", "giày công sở": "dress shoes", "ủng": "boots",
    "giày lười": "slip-ons", "giày cao gót": "heels", # Loại giày

    "túi đeo chéo": "crossbody bag", "ba lô": "backpack", "túi xách tay": "hand bag", 
    "túi đeo vai": "shoulder bag", "túi vải vuông to": "tote", "túi nilon": "plastic bag", "cặp sách (có thể có dây đeo)": "messenger bag", # Loại túi

    "đồng hồ đeo tay": "wristwatch", "vòng tay": "bracelet", "vòng cổ": "necklace", "vòng chân": "anklet", 
    "ví vuông": "clutch", "ví da": "wallet", "kính": "glasses", "thẻ nhân viên": "ID card", "khăn quàng cổ": "scarf",
    "khẩu trang": "mask", "sách": "book", "giấy": "paper", "điện thoại": "phone", "ô": "umbrella", "vali": "suitcase",
    "chổi lau": "mop", "cái xô": "bucket"   # Khác
}

# Các nhóm thuộc tính
attribute_groups = {
    "Thông tin cơ bản": {
        "Giới tính": ["nam", "nữ", "không rõ"],
        "Tuổi": ["20s", "1-5", "5-10", "10s", "30s", "40s", "50s", "60s", "70s", "80s", "không rõ"],
        "Vóc dáng": ["gầy", "bình thường", "mập"],
    },
    "Tóc": {
        "Màu": ["đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ", "không có"],
        "Độ dài": ["ngắn", "trung bình", "dài", "không rõ", "không có"],
    },
    "Mũ": {
        "Màu": ["không có","đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ"],
        "Loại": ["không có", "mũ bảo hiểm", "mũ công nhân", "mũ lưỡi trai", "nón lá", "mũ tai bèo", "mũ nồi", "không rõ"]
    },
    "Áo": {
        "Màu":["trắng", "đen", "nâu", "vàng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ", "không có"],
        "Loại": ["áo phông", "áo sơ mi", "áo khoác", "áo phông nữ", "áo vest", "váy", "áo polo", "áo khoác dày", "măng tô", "hoodie"],
        "Họa tiết": ["trơn", "sọc ngang", "sọc dọc", "in chữ", "in hình", "logo", "kẻ caro", "đốm", "màu sặc sỡ", "có vùng khác màu"],
        "Tay áo": ["ngắn tay", "dài tay", "không rõ"]
    },
    "Quần": {
        "Màu": ["đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ"],
        "Loại": ["quần âu", "quần bò", "quần short/quần đùi", "quần thô túi hộp", "quần vải bó chân", "quần vải chỉ qua đầu gối", "quần vải thể thao", "chân váy", "váy", "quần cạp cao có dây đeo vai"],
        "Họa tiết": ["trơn", "sọc ngang", "sọc dọc", "in chữ", "in hình", "logo", "kẻ caro", "đốm", "màu sặc sỡ", "có vùng khác màu"],
        "Chiều dài": ["dài", "ngắn", "không rõ"]
    },
    "Giày": {
        "Màu": ["đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ"],
        "Loại": ["không rõ", "giày thể thao", "dép", "tông", "giày công sở", "ủng", "giày lười", "giày cao gót"]
    },
    "Túi": {
        "Màu": ["đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím", "không rõ", "không có"],
        "Loại": ["không có", "ba lô", "túi đeo chéo", "túi xách tay", "túi đeo vai", "túi vải vuông to", "túi nilon", "cặp sách (có thể có dây đeo)"]
    },
    "Khác": {
        "Màu": ["không có", "đen", "nâu", "vàng", "trắng", "xanh", "đỏ", "xám", "xanh lá", "hồng", "tím"],
        "Loại": ["không có", "khẩu trang", "điện thoại", "ô", "vali", "chổi lau", "cái xô", "đồng hồ đeo tay", "vòng tay", "vòng cổ", "vòng chân", "ví vuông", "ví da", "kính", "thẻ nhân viên", "khăn quàng cổ", "sách", "giấy"]
    }
}

class LabelingTool:
    def __init__(self, root):
        self.root = root
        try:
            self.root.attributes('-zoomed', True)
        except:
            self.root.geometry(f'{root.winfo_screenwidth()}x{root.winfo_screenheight()}')
        self.root.title("Công cụ gán nhãn ảnh")

        self.main_frame = tk.Frame(root)
        self.main_frame.pack(fill="both", expand=True)

        # === Khung ảnh ===
        self.image_frame = tk.Frame(self.main_frame)
        self.image_frame.pack(side="left", padx=20, pady=10, anchor="n")

        self.image_panel = tk.Label(self.image_frame, width=700, height=700, bg="gray")
        self.image_panel.pack()

        self.nav_frame = tk.Frame(self.image_frame)
        self.nav_frame.pack(pady=10)

        self.prev_button = tk.Button(self.nav_frame, text="← Ảnh trước", command=self.prev_image, font=("Roboto", 14))
        self.prev_button.pack(side="left", padx=10)

        self.next_button = tk.Button(self.nav_frame, text="Ảnh tiếp →", command=self.next_image, font=("Roboto", 14))
        self.next_button.pack(side="left", padx=10)
        
        # Thêm label hiển thị tên ảnh
        self.filename_label = tk.Label(self.image_frame, text="", font=("Roboto", 14))
        self.filename_label.pack(pady=5)
        
        # Thêm frame cho chức năng tìm kiếm ảnh
        self.search_frame = tk.Frame(self.image_frame)
        self.search_frame.pack(pady=10)
        
        # Thêm ô nhập tên ảnh
        self.search_entry = tk.Entry(self.search_frame, width=30, font=("Roboto", 14))
        self.search_entry.pack(side="left", padx=5)
        
        # Thêm nút tìm kiếm
        self.search_button = tk.Button(self.search_frame, text="Tìm ảnh", command=self.find_image, font=("Roboto", 14))
        self.search_button.pack(side="left", padx=5)

        # === Style cho Scrollbar ===
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Vertical.TScrollbar", 
                        width=25,
                        background="#838683",
                        troughcolor="#f0f0f0",
                        arrowcolor="black",
                        borderwidth=2,
                        relief="raised")
        
        style.map("Vertical.TScrollbar",
                 background=[('active', '#45a049'), ('pressed', '#367c39')],
                 arrowcolor=[('active', 'white'), ('pressed', 'white')])

        # === Khung thuộc tính với thanh cuộn ===
        self.control_outer_frame = tk.Frame(self.main_frame)
        self.control_outer_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        self.canvas = tk.Canvas(self.control_outer_frame)
        self.scrollbar = ttk.Scrollbar(self.control_outer_frame, orient="vertical", command=self.canvas.yview, style="Vertical.TScrollbar")
        self.scrollable_frame = tk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Thêm binding cho sự kiện lăn chuột
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.control_frame = self.scrollable_frame  # gán để dùng chung

        self.comboboxes = {}
        self.dynamic_fields = {"Túi": [], "Khác": []}
        
        # Tạo fonts với Roboto
        larger_font = ("Roboto", 16)
        larger_font_bold = ("Roboto", 16, "bold")

        # Keep track of group frames for dynamic groups
        self.group_frames = {}

        # Tạo style cho separator
        style = ttk.Style()
        style.configure("TSeparator", background="#cccccc")

        for i, (group_name, group_fields) in enumerate(attribute_groups.items()):
            if i > 0:
                separator = ttk.Separator(self.control_frame, orient="horizontal")
                separator.pack(fill="x", pady=10, padx=5)

            group_frame = tk.Frame(self.control_frame, bd=2, relief="groove", padx=10, pady=10)
            group_frame.pack(fill="x", pady=(5, 10))
            
            group_header_frame = tk.Frame(group_frame)
            group_header_frame.pack(fill="x")
            
            group_label = tk.Label(group_header_frame, text=group_name, font=larger_font_bold)
            group_label.pack(side="left", anchor="w")
            
            if group_name in ["Túi", "Khác"]:
                self.group_frames[group_name] = group_frame
                add_btn = tk.Button(
                    group_header_frame, 
                    text="+", 
                    command=lambda g=group_name: self.add_group_attributes(g),
                    width=2,
                    font=larger_font
                )
                # Thay đổi từ side="right" thành side="left" và thêm padx sau group_label
                add_btn.pack(side="left", padx=10)

            for attr, options in group_fields.items():
                attr_frame = tk.Frame(group_frame)
                attr_frame.pack(fill="x", pady=5)

                label = tk.Label(attr_frame, text=attr + ":", font=larger_font, width=20, anchor="w")
                label.pack(side="left")

                cb = ttk.Combobox(attr_frame, values=options, state="readonly", font=larger_font, width=20)  # Giảm từ 30 xuống 20
                cb.set(options[0])  # Thiết lập mặc định là lựa chọn đầu tiên
                cb.pack(side="left")

                key = f"{group_name}:{attr}"
                self.comboboxes[key] = [cb]

        # Thêm ô ghi chú với font Roboto
        note_label = tk.Label(self.control_frame, text="Lưu ý:", font=("Roboto", 14))
        note_label.pack(anchor="w", pady=(15, 0))
        self.note_entry = tk.Text(self.control_frame, height=3, font=("Roboto", 14), wrap="word")
        self.note_entry.pack(fill="x", padx=5)

        # Nút lưu với font Roboto
        self.save_button = tk.Button(self.control_frame, text="Lưu nhãn", command=self.save_label, 
                                   font=("Roboto", 14), bg="#4CAF50", fg="white")
        self.save_button.pack(pady=20)

        # Lưu trữ label cho mỗi ảnh
        self.image_labels = {}
        
        # Load ảnh từ cấu trúc thư mục phức tạp
        self.folder = filedialog.askdirectory(title="Chọn thư mục ảnh")
        self.load_images_from_subfolders()
        
        self.excel_file = "labels.xlsx"
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Tên ảnh", "Miêu tả", "Lưu ý"])
            wb.save(self.excel_file)

        # Delay loading the first image until the UI is ready
        self.root.after(100, self.load_image)

    def load_images_from_subfolders(self):
        """Tải danh sách ảnh từ tất cả các thư mục con"""
        self.image_list = []
        self.full_paths = []
        
        # Tìm tất cả các thư mục con cấp 1
        subfolders = [f for f in os.listdir(self.folder) if os.path.isdir(os.path.join(self.folder, f))]
        
        # Sắp xếp thư mục theo thứ tự tự nhiên
        import re
        def natural_sort_key(s):
            return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]
        
        subfolders.sort(key=natural_sort_key)
        
        # Duyệt qua từng thư mục con và lấy danh sách ảnh
        for subfolder in subfolders:
            subfolder_path = os.path.join(self.folder, subfolder)
            
            # Lấy tất cả ảnh từ thư mục con này
            image_extensions = ['*.jpg', '*.jpeg', '*.png']
            for ext in image_extensions:
                image_paths = glob.glob(os.path.join(subfolder_path, ext))
                
                for img_path in image_paths:
                    # Lưu đường dẫn đầy đủ đến ảnh
                    self.full_paths.append(img_path)
                    
                    # Tạo tên hiển thị: "tên_folder - tên_ảnh"
                    img_filename = os.path.basename(img_path)
                    display_name = f"{subfolder} - {img_filename}"
                    self.image_list.append(display_name)
        
        # Sắp xếp ảnh theo thứ tự tự nhiên
        combined = list(zip(self.image_list, self.full_paths))
        combined.sort(key=lambda x: natural_sort_key(x[0]))
        
        if combined:
            self.image_list, self.full_paths = zip(*combined)
            self.image_list = list(self.image_list)
            self.full_paths = list(self.full_paths)
        
        self.image_index = 0
    
    def save_current_labels(self):
        """Lưu lại trạng thái hiện tại của tất cả combobox"""
        if not self.image_list:
            return
        
        current_image = self.image_list[self.image_index]
        label_state = {}
        
        # Lưu trạng thái của tất cả combobox
        for key, cb_list in self.comboboxes.items():
            # Lọc các combobox còn tồn tại và lấy giá trị
            valid_values = []
            for cb in cb_list:
                try:
                    if cb.winfo_exists():
                        valid_values.append(cb.get())
                except (tk.TclError, AttributeError):
                    # Bỏ qua các widget không tồn tại
                    continue
                
            if valid_values:  # Chỉ lưu nếu có giá trị hợp lệ
                label_state[key] = valid_values
        
        # Lưu note
        note_text = self.note_entry.get("1.0", "end").strip()
        if note_text:
            label_state["note"] = note_text
            
        # Lưu vào từ điển
        self.image_labels[current_image] = label_state

    def restore_labels(self):
        """Khôi phục lại trạng thái đã lưu của các combobox"""
        if not self.image_list:
            return
            
        current_image = self.image_list[self.image_index]
        
        # Kiểm tra xem có dữ liệu đã lưu không
        if current_image not in self.image_labels:
            return
            
        label_state = self.image_labels[current_image]
        
        # Khôi phục trạng thái cho các combobox
        for key, values in label_state.items():
            if key == "note":
                # Khôi phục note
                self.note_entry.delete("1.0", "end")
                self.note_entry.insert("1.0", values)
            elif key in self.comboboxes:
                try:
                    # Khôi phục combobox
                    cb_list = self.comboboxes[key]
                    
                    # Thêm combobox mới nếu cần
                    while len(cb_list) < len(values):
                        group, attr = key.split(':')
                        self.add_group_attributes(group)
                        cb_list = self.comboboxes[key]
                    
                    # Đặt giá trị
                    for i, value in enumerate(values):
                        if i < len(cb_list) and cb_list[i].winfo_exists():
                            cb_list[i].set(value)
                except (tk.TclError, AttributeError) as e:
                    # Bỏ qua lỗi nếu widget không tồn tại
                    print(f"Không thể khôi phục {key}: {e}")
                    continue

    def resize_image(self, image, max_w, max_h):
        w, h = image.size
        scale = min(max_w / w, max_h / h)
        return image.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)

    def load_image(self):
        if not self.image_list:
            messagebox.showinfo("Thông báo", "Không tìm thấy ảnh nào trong thư mục.")
            return
            
        # Lưu lại label của ảnh hiện tại trước khi chuyển sang ảnh khác
        self.save_current_labels()
        
        # Xóa tất cả các combobox phụ đã thêm vào trước khi load ảnh mới
        self.reset_dynamic_fields()
        
        # Kiểm tra và sửa lỗi thiếu combobox
        self.check_combobox_existence()
        
        # Tải ảnh mới
        img_path = self.full_paths[self.image_index]
        image = Image.open(img_path)
        image = self.resize_image(image, 700, 700)

        bg = Image.new("RGB", (700, 700), "gray")
        offset_x = (700 - image.width) // 2
        offset_y = (700 - image.height) // 2
        bg.paste(image, (offset_x, offset_y))

        self.tk_image = ImageTk.PhotoImage(bg)
        self.image_panel.config(image=self.tk_image)
        self.image_panel.image = self.tk_image
        
        # Cập nhật tên ảnh hiển thị
        self.filename_label.config(text=f"Tên ảnh: {self.image_list[self.image_index]}")
        
        # Khôi phục lại label đã lưu (nếu có)
        self.restore_labels()

    def reset_dynamic_fields(self):
        """Xóa tất cả các trường phụ đã thêm vào cho nhóm Túi và Khác"""
        for group_name in ["Túi", "Khác"]:
            if group_name in self.group_frames:
                # 1. Lưu lại frame header (nút + và tên nhóm)
                header_frame = None
                for child in self.group_frames[group_name].winfo_children():
                    if isinstance(child, tk.Frame) and len(child.winfo_children()) > 0:
                        if isinstance(child.winfo_children()[0], tk.Label) and child.winfo_children()[0].cget("text") == group_name:
                            header_frame = child
                            break
            
                # 2. Xóa tất cả widget con trừ header
                for child in list(self.group_frames[group_name].winfo_children()):
                    if child != header_frame:
                        child.destroy()
            
                # 3. Tạo lại frame cho các combobox đầu tiên
                self._create_default_attributes(group_name)
            
                # 4. Reset dictionary comboboxes
                for key in list(self.comboboxes.keys()):
                    if key.startswith(f"{group_name}:"):
                        attr = key.split(':')[1]
                        # Tìm combobox mới tạo cho thuộc tính này
                        for child in self.group_frames[group_name].winfo_children():
                            if isinstance(child, tk.Frame):
                                for widget in child.winfo_children():
                                    if isinstance(widget, tk.Label) and widget.cget("text") == f"{attr}:":
                                        # Tìm combobox bên cạnh label này
                                        for w in child.winfo_children():
                                            if isinstance(w, ttk.Combobox):
                                                self.comboboxes[key] = [w]
                                                break

    def _create_default_attributes(self, group_name):
        """Tạo lại các thuộc tính mặc định cho một nhóm"""
        attributes = attribute_groups[group_name]
        
        for attr, options in attributes.items():
            attr_frame = tk.Frame(self.group_frames[group_name])
            attr_frame.pack(fill="x", pady=5)

            label = tk.Label(attr_frame, text=attr + ":", font=("Roboto", 16), width=20, anchor="w")
            label.pack(side="left")

            cb = ttk.Combobox(attr_frame, values=options, state="readonly", font=("Roboto", 16), width=20)  # Giảm từ 30 xuống 20
            cb.set(options[0])  # Thiết lập mặc định là lựa chọn đầu tiên
            cb.pack(side="left")

            key = f"{group_name}:{attr}"
            self.comboboxes[key] = [cb]

    def get_labels(self):
        # Tạo từ điển để lưu các thuộc tính theo nhóm
        grouped_attributes = {
            "gender": [],      # Giới tính
            "age": [],         # Tuổi
            "vitality": [],    # Vóc dáng
            "hair": [],        # Tóc (màu, độ dài)
            "hat": [],         # Mũ (màu, loại)
            "shirt": [],       # Áo (màu, loại, họa tiết, tay áo)
            "pants": [],       # Quần (màu, loại, họa tiết, chiều dài)
            "shoes": [],       # Giày (màu, loại)
            "bag": [],         # Túi (màu, loại)
            "other": []        # Khác (màu, loại)
        }
        
        # Ánh xạ tên nhóm trong tiếng Việt sang tiếng Anh
        group_mapping = {
            "Thông tin cơ bản": {"Giới tính": "gender", "Tuổi": "age", "Vóc dáng": "vitality"},
            "Tóc": {"Màu": "hair", "Độ dài": "hair"},
            "Mũ": {"Màu": "hat", "Loại": "hat"},
            "Áo": {"Màu": "shirt", "Loại": "shirt", "Họa tiết": "shirt", "Tay áo": "shirt"},
            "Quần": {"Màu": "pants", "Loại": "pants", "Họa tiết": "pants", "Chiều dài": "pants"},
            "Giày": {"Màu": "shoes", "Loại": "shoes"},
            "Túi": {"Màu": "bag", "Loại": "bag"},
            "Khác": {"Màu": "other", "Loại": "other"}
        }
        
        # Thu thập các giá trị từ combobox và lưu trữ riêng cho túi và khác
        bag_entries = []  # Lưu trữ các bộ [màu, loại] cho túi
        other_entries = []  # Lưu trữ các bộ [màu, loại] cho các vật dụng khác
        
        # Thu thập các giá trị từ combobox
        for key, cb_list in self.comboboxes.items():
            group, attr = key.split(':')
            
            for i, cb in enumerate(cb_list):
                try:
                    # Kiểm tra xem widget còn tồn tại không
                    if not cb.winfo_exists():
                        continue
                        
                    val = cb.get()
                    if val in ["không rõ", "không có"]:
                        val_en = "unknown" if val == "không rõ" else "none"
                    else:
                        val_en = translation_dict.get(val, val)
                    
                    # Xác định nhóm tiếng Anh tương ứng
                    eng_group = group_mapping[group][attr]
                    
                    # Xử lý đặc biệt cho túi và khác
                    if group == "Túi":
                        # Túi thứ i
                        if i >= len(bag_entries):
                            bag_entries.append({"màu": None, "loại": None})
                        
                        if attr == "Màu":
                            bag_entries[i]["màu"] = val_en
                        elif attr == "Loại":
                            bag_entries[i]["loại"] = val_en
                    elif group == "Khác":
                        # Vật dụng khác thứ i
                        if i >= len(other_entries):
                            other_entries.append({"màu": None, "loại": None})
                        
                        if attr == "Màu":
                            other_entries[i]["màu"] = val_en
                        elif attr == "Loại":
                            other_entries[i]["loại"] = val_en
                    else:
                        # Các nhóm còn lại xử lý bình thường
                        grouped_attributes[eng_group].append(val_en)
                except (tk.TclError, AttributeError) as e:
                    # Bỏ qua lỗi nếu widget không tồn tại
                    print(f"Lỗi khi lấy giá trị {key} thứ {i}: {e}")
                    continue
        
        # Tạo chuỗi kết quả theo định dạng yêu cầu
        result = []
        
        # Xử lý gender và age (thông tin cơ bản) trước
        if grouped_attributes["gender"]:
            result.append(f"gender: {grouped_attributes['gender'][0]}")
        
        if grouped_attributes["age"]:
            result.append(f"age: {grouped_attributes['age'][0]}")

        if grouped_attributes["vitality"]:
            result.append(f"vitality: {grouped_attributes['vitality'][0]}")
        
        # Xử lý các nhóm thông thường (không phải túi và khác)
        for group in ["hair", "hat", "shirt", "pants", "shoes"]:
            if grouped_attributes[group]:
                values = list(dict.fromkeys(grouped_attributes[group]))
                
                # Nếu tất cả là 'none' hoặc 'unknown', chỉ giữ lại một giá trị
                if all(v in ["none", "unknown"] for v in values) and len(values) > 1:
                    values = [values[0]]
                    
                # Tạo chuỗi định dạng (value1, value2, ...)
                if len(values) > 0:
                    formatted_values = f"({', '.join(values)})"
                    result.append(f"{group}: {formatted_values}")
        
        # Xử lý túi - mỗi túi được miêu tả trong một ngoặc riêng
        if any(entry["màu"] not in [None, "none"] or entry["loại"] not in [None, "none"] for entry in bag_entries):
            bag_parts = []
            
            for entry in bag_entries:
                values = []
                if entry["màu"]:
                    values.append(entry["màu"])
                if entry["loại"]:
                    values.append(entry["loại"])
                
                if values and not (len(values) == 1):
                    bag_parts.append(f"({', '.join(values)})")
            
            if bag_parts:
                result.append(f"bag: {' '.join(bag_parts)}")
        
        # Xử lý khác - mỗi vật dụng được miêu tả trong một ngoặc riêng
        if any(entry["màu"] not in [None, "none"] or entry["loại"] not in [None, "none"] for entry in other_entries):
            other_parts = []
            
            for entry in other_entries:
                values = []
                if entry["màu"] not in [None, "none"]:
                    values.append(entry["màu"])
                if entry["loại"] not in [None, "none"]:
                    values.append(entry["loại"])
                
                if values and not (len(values) == 1 and values[0] == "unknown"):
                    other_parts.append(f"({', '.join(values)})")
            
            if other_parts:
                result.append(f"other: {' '.join(other_parts)}")
        
        # Kết hợp tất cả các nhóm thành một chuỗi duy nhất
        return ", ".join(result)

    def save_label(self):
        description = self.get_labels()
        note_text = self.note_entry.get("1.0", "end").strip()
        filename = self.image_list[self.image_index]
        
        # Lưu lại label hiện tại vào bộ nhớ
        self.save_current_labels()

        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([filename, description, note_text])
        wb.save(self.excel_file)
        
        # Thay đổi nút lưu để hiển thị dấu tích - giữ trạng thái này cho đến khi chuyển ảnh
        self.save_button.config(text="✓ Đã lưu")
        self.save_button.config(bg="#45a049")  # Làm xanh đậm nút khi lưu thành công

    def prev_image(self):
        if self.image_index > 0:
            # Đổi trạng thái nút lưu về ban đầu khi chuyển ảnh
            self.save_button.config(text="Lưu nhãn", bg="#4CAF50")
            
            self.image_index -= 1
            self.load_image()
            
    def next_image(self):
        if self.image_index < len(self.image_list) - 1:
            # Đổi trạng thái nút lưu về ban đầu khi chuyển ảnh
            self.save_button.config(text="Lưu nhãn", bg="#4CAF50")
            
            self.image_index += 1
            self.load_image()
    
    def find_image(self):
        search_term = self.search_entry.get().strip().lower()
        if not search_term:
            return
            
        # Đầu tiên, thử tìm folder theo số
        if search_term.isdigit():
            folder_matches = [i for i, img_name in enumerate(self.image_list) 
                         if img_name.startswith(f"{search_term} -")]
        
            if folder_matches:
                # Đổi trạng thái nút lưu về ban đầu khi tìm và chuyển ảnh
                self.save_button.config(text="Lưu nhãn", bg="#4CAF50")
                
                # Lưu lại label của ảnh hiện tại trước khi chuyển sang ảnh khác
                self.save_current_labels()
                
                # Chọn ảnh đầu tiên trong folder có số này
                self.image_index = folder_matches[0]
                self.load_image()
                
                # Hiển thị thông báo về số lượng ảnh tìm thấy
                if len(folder_matches) > 1:
                    message = f"Tìm thấy {len(folder_matches)} ảnh trong folder {search_term}. Hiển thị ảnh đầu tiên."
                    tk.messagebox.showinfo("Kết quả tìm kiếm", message)
                return
        
        # Nếu không phải số hoặc không tìm thấy folder, tìm theo cách thông thường
        matching_images = [i for i, img_name in enumerate(self.image_list) 
                         if search_term in img_name.lower()]
        
        if matching_images:
            # Đổi trạng thái nút lưu về ban đầu khi tìm và chuyển ảnh
            self.save_button.config(text="Lưu nhãn", bg="#4CAF50")
            
            # Lưu lại label của ảnh hiện tại trước khi chuyển sang ảnh khác
            self.save_current_labels()
            
            # Chọn ảnh đầu tiên phù hợp
            self.image_index = matching_images[0]
            self.load_image()
            
            # Hiển thị thông báo về số lượng ảnh tìm thấy
            if len(matching_images) > 1:
                message = f"Tìm thấy {len(matching_images)} ảnh. Hiển thị ảnh đầu tiên."
                tk.messagebox.showinfo("Kết quả tìm kiếm", message)
        else:
            tk.messagebox.showinfo("Không tìm thấy", f"Không tìm thấy ảnh nào khớp với '{search_term}'")

    def _on_mousewheel(self, event):
        # Xác định hệ điều hành và xử lý sự kiện lăn chuột tương ứng
        if event.num == 4:  # Linux scroll up
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:  # Linux scroll down
            self.canvas.yview_scroll(1, "units")
        else:  # Windows
            # Với Windows, event.delta có thể là dương hoặc âm tùy thuộc vào hướng lăn
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def add_group_attributes(self, group_name):
        """Add a complete set of attribute boxes for the specified group"""
        # Get the attributes for this group
        attributes = attribute_groups[group_name]
    
        # Create a new frame to hold the duplicated attributes horizontally
        new_row_frame = tk.Frame(self.group_frames[group_name])
        new_row_frame.pack(fill="x", pady=5)
    
        # Add each attribute in the group
        for attr, options in attributes.items():
            # Create a frame for this attribute
            attr_frame = tk.Frame(new_row_frame)
            attr_frame.pack(side="left", padx=10)
            
            # Label với font Roboto
            label = tk.Label(attr_frame, text=attr + ":", font=("Roboto", 16), anchor="w")
            label.pack(anchor="w")
            
            # Combobox với font Roboto
            cb = ttk.Combobox(attr_frame, values=options, state="readonly", font=("Roboto", 16), width=15)
            cb.set(options[0])  # Thiết lập mặc định là lựa chọn đầu tiên
            cb.pack()
            
            # Add to our comboboxes dictionary
            key = f"{group_name}:{attr}"
            self.comboboxes.setdefault(key, []).append(cb)

    def check_combobox_existence(self):
        """Kiểm tra xem tất cả các combobox cần thiết có tồn tại không và tạo lại nếu cần"""
        for group_name, group_fields in attribute_groups.items():
            for attr in group_fields:
                key = f"{group_name}:{attr}"
                if key not in self.comboboxes or not self.comboboxes[key]:
                    print(f"Phát hiện thiếu combobox: {key}")
                    # Tìm frame chứa thuộc tính
                    if group_name in self.group_frames:
                        attr_frame = None
                        # Tìm frame con đầu tiên (bỏ qua header)
                        children = self.group_frames[group_name].winfo_children()
                        if len(children) > 1:
                            for child in children[1:]:
                                if isinstance(child, tk.Frame):
                                    attr_frame = child
                                    break
                    
                        if attr_frame is None:
                            # Tạo frame mới nếu không tìm thấy
                            attr_frame = tk.Frame(self.group_frames[group_name])
                            attr_frame.pack(fill="x", pady=5)
                        
                        # Tạo label và combobox mới
                        options = group_fields[attr]
                        label = tk.Label(attr_frame, text=attr + ":", font=("Roboto", 16), width=20, anchor="w")
                        label.pack(side="left")
                        
                        cb = ttk.Combobox(attr_frame, values=options, state="readonly", font=("Roboto", 16), width=20)  # Giảm từ 30 xuống 20
                        cb.set(options[0])
                        cb.pack(side="left")
                        
                        # Cập nhật dictionary
                        self.comboboxes[key] = [cb]

# Chạy ứng dụng
if __name__ == "__main__":
    root = tk.Tk()
    app = LabelingTool(root)
    root.mainloop()